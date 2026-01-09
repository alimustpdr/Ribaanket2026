#!/usr/bin/env python3
"""
Survey Extraction Tool for Ribaanket2026 Project

This script extracts content from preschool survey documents:
- PDF forms → Plain text (.txt)
- XLSX result spreadsheets → CSV files (per sheet) + JSON metadata
"""

import argparse
import json
import os
import sys
from pathlib import Path
from typing import List, Dict, Any

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install -r requirements.txt")
    sys.exit(1)


# Target preschool files to process
TARGET_PDF_FILES = [
    "18115003_RYBA_FORM_Veli_Okul_Oncesi.pdf",
    "18174526_RYBA_FORM_OYretmen_Okul_Oncesi.pdf"
]

TARGET_XLSX_FILES = [
    "26134246_ribaokuloncesiokulsonuccizelgesi.xlsx",
    "26134300_ribaokuloncesisinifsonuccizelgesi (1).xlsx"
]


def find_file_case_insensitive(directory: Path, target_filename: str) -> Path:
    """
    Find a file in directory matching target filename (case-insensitive, handles spaces).
    
    Args:
        directory: Directory to search in
        target_filename: Target filename to find
        
    Returns:
        Path to the found file
        
    Raises:
        FileNotFoundError: If file not found
    """
    # Normalize the target filename for comparison
    target_normalized = target_filename.lower().replace(" ", "").replace("%20", "")
    
    for file_path in directory.iterdir():
        if file_path.is_file():
            # Normalize the current filename for comparison
            current_normalized = file_path.name.lower().replace(" ", "").replace("%20", "")
            if current_normalized == target_normalized:
                return file_path
    
    # If exact match not found, try matching by the base ID number
    target_id = target_filename.split("_")[0] if "_" in target_filename else None
    if target_id:
        for file_path in directory.iterdir():
            if file_path.is_file() and file_path.name.startswith(target_id):
                # Check if extension matches
                if target_filename.endswith(".pdf") and file_path.suffix.lower() == ".pdf":
                    return file_path
                elif target_filename.endswith(".xlsx") and file_path.suffix.lower() == ".xlsx":
                    return file_path
    
    raise FileNotFoundError(f"File not found: {target_filename}")


def extract_pdf_text(pdf_path: Path, output_dir: Path) -> None:
    """
    Extract text content from PDF file and save to .txt file.
    
    Args:
        pdf_path: Path to PDF file
        output_dir: Directory to save output text file
    """
    print(f"  Extracting text from: {pdf_path.name}")
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract text from all pages
            all_text = []
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text:
                    all_text.append(f"=== Page {i + 1} ===\n{text}\n")
            
            # Save to output file
            output_filename = pdf_path.stem + ".txt"
            output_path = output_dir / output_filename
            
            with open(output_path, "w", encoding="utf-8") as f:
                f.write("\n".join(all_text))
            
            print(f"    → Saved to: {output_path}")
            print(f"    → Pages extracted: {len(pdf.pages)}")
            
    except Exception as e:
        print(f"    ERROR: Failed to extract PDF: {e}")


def get_cell_type(cell) -> str:
    """Determine the type of a cell value."""
    if cell.value is None:
        return "empty"
    elif cell.data_type == "f":
        return "formula"
    elif cell.data_type == "n":
        return "number"
    elif cell.data_type == "b":
        return "boolean"
    elif cell.data_type == "s":
        return "string"
    elif cell.data_type == "d":
        return "date"
    else:
        return "other"


def extract_xlsx_to_csv(xlsx_path: Path, output_dir: Path) -> None:
    """
    Extract XLSX file sheets to individual CSV files.
    
    Args:
        xlsx_path: Path to XLSX file
        output_dir: Directory to save CSV files
    """
    print(f"  Converting to CSV: {xlsx_path.name}")
    
    try:
        workbook = load_workbook(xlsx_path, data_only=True)
        base_filename = xlsx_path.stem
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Create safe filename from sheet name
            safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_sheet_name = safe_sheet_name.replace(" ", "_")
            if not safe_sheet_name:
                safe_sheet_name = f"sheet_{workbook.sheetnames.index(sheet_name)}"
            
            csv_filename = f"{base_filename}_{safe_sheet_name}.csv"
            csv_path = output_dir / csv_filename
            
            # Write CSV
            with open(csv_path, "w", encoding="utf-8") as f:
                for row in sheet.iter_rows(values_only=True):
                    # Convert row values to strings and handle None
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    # Escape quotes and handle commas
                    row_data = [f'"{cell}"' if ',' in cell or '"' in cell else cell for cell in row_data]
                    f.write(",".join(row_data) + "\n")
            
            print(f"    → Saved sheet '{sheet_name}' to: {csv_path}")
        
        workbook.close()
        
    except Exception as e:
        print(f"    ERROR: Failed to convert to CSV: {e}")


def extract_xlsx_metadata(xlsx_path: Path, output_dir: Path) -> None:
    """
    Extract XLSX metadata and structure to JSON file.
    
    Args:
        xlsx_path: Path to XLSX file
        output_dir: Directory to save JSON metadata
    """
    print(f"  Extracting metadata: {xlsx_path.name}")
    
    try:
        workbook = load_workbook(xlsx_path, data_only=False)
        metadata = {
            "filename": xlsx_path.name,
            "sheets": []
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Analyze cell types
            cell_types = {}
            formulas = []
            
            for row in sheet.iter_rows(max_row=min(max_row, 1000)):  # Limit analysis to first 1000 rows
                for cell in row:
                    cell_type = get_cell_type(cell)
                    cell_types[cell_type] = cell_types.get(cell_type, 0) + 1
                    
                    # Collect formulas
                    if cell.data_type == "f" and cell.value:
                        formulas.append({
                            "cell": cell.coordinate,
                            "formula": cell.value
                        })
            
            # Try to detect header row (first non-empty row with mostly strings)
            header_row = None
            for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
                non_empty = [cell for cell in row if cell is not None]
                if len(non_empty) > max_col * 0.5:  # More than 50% filled
                    header_row = i
                    break
            
            sheet_metadata = {
                "name": sheet_name,
                "dimensions": {
                    "rows": max_row,
                    "columns": max_col
                },
                "header_row": header_row,
                "cell_types": cell_types,
                "formula_count": len(formulas),
                "formulas_sample": formulas[:10]  # First 10 formulas
            }
            
            metadata["sheets"].append(sheet_metadata)
        
        # Save metadata to JSON
        json_filename = xlsx_path.stem + "_metadata.json"
        json_path = output_dir / json_filename
        
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)
        
        print(f"    → Saved metadata to: {json_path}")
        print(f"    → Sheets analyzed: {len(metadata['sheets'])}")
        
        workbook.close()
        
    except Exception as e:
        print(f"    ERROR: Failed to extract metadata: {e}")


def process_pdf_files(input_dir: Path, output_dir: Path, target_files: List[str]) -> int:
    """
    Process PDF files and extract text.
    
    Returns:
        Number of files successfully processed
    """
    print("\n" + "="*60)
    print("Processing PDF files...")
    print("="*60)
    
    processed = 0
    for target_file in target_files:
        try:
            pdf_path = find_file_case_insensitive(input_dir, target_file)
            extract_pdf_text(pdf_path, output_dir)
            processed += 1
        except FileNotFoundError as e:
            print(f"  WARNING: {e}")
        except Exception as e:
            print(f"  ERROR processing {target_file}: {e}")
    
    return processed


def process_xlsx_files(input_dir: Path, csv_output_dir: Path, json_output_dir: Path, target_files: List[str]) -> int:
    """
    Process XLSX files and extract to CSV and JSON.
    
    Returns:
        Number of files successfully processed
    """
    print("\n" + "="*60)
    print("Processing XLSX files...")
    print("="*60)
    
    processed = 0
    for target_file in target_files:
        try:
            xlsx_path = find_file_case_insensitive(input_dir, target_file)
            print(f"\nProcessing: {xlsx_path.name}")
            extract_xlsx_to_csv(xlsx_path, csv_output_dir)
            extract_xlsx_metadata(xlsx_path, json_output_dir)
            processed += 1
        except FileNotFoundError as e:
            print(f"  WARNING: {e}")
        except Exception as e:
            print(f"  ERROR processing {target_file}: {e}")
    
    return processed


def main():
    """Main entry point for the extraction tool."""
    parser = argparse.ArgumentParser(
        description="Extract content from preschool survey documents (PDF and XLSX files)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process all files (default)
  python scripts/extract_surveys.py
  
  # Process only PDF files
  python scripts/extract_surveys.py --pdf
  
  # Process only XLSX files
  python scripts/extract_surveys.py --xlsx
  
  # Custom input/output directories
  python scripts/extract_surveys.py --input-dir /path/to/files --output-dir /path/to/output
        """
    )
    
    parser.add_argument(
        "--input-dir",
        type=str,
        default=".",
        help="Input directory containing survey files (default: current directory)"
    )
    
    parser.add_argument(
        "--output-dir",
        type=str,
        default="extracted/okuloncesi",
        help="Output directory for extracted files (default: extracted/okuloncesi)"
    )
    
    parser.add_argument(
        "--pdf",
        action="store_true",
        help="Process only PDF files"
    )
    
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Process only XLSX files"
    )
    
    args = parser.parse_args()
    
    # Convert paths to Path objects
    input_dir = Path(args.input_dir).resolve()
    output_base_dir = Path(args.output_dir).resolve()
    
    # Validate input directory
    if not input_dir.exists():
        print(f"ERROR: Input directory does not exist: {input_dir}")
        sys.exit(1)
    
    # Create output directories
    pdf_output_dir = output_base_dir / "pdf_text"
    csv_output_dir = output_base_dir / "xlsx_csv"
    json_output_dir = output_base_dir / "xlsx_json"
    
    pdf_output_dir.mkdir(parents=True, exist_ok=True)
    csv_output_dir.mkdir(parents=True, exist_ok=True)
    json_output_dir.mkdir(parents=True, exist_ok=True)
    
    print("="*60)
    print("Survey Extraction Tool")
    print("="*60)
    print(f"Input directory:  {input_dir}")
    print(f"Output directory: {output_base_dir}")
    print("="*60)
    
    # Determine what to process
    process_pdf = args.pdf or not (args.pdf or args.xlsx)
    process_xlsx = args.xlsx or not (args.pdf or args.xlsx)
    
    total_processed = 0
    
    # Process PDF files
    if process_pdf:
        pdf_count = process_pdf_files(input_dir, pdf_output_dir, TARGET_PDF_FILES)
        total_processed += pdf_count
    
    # Process XLSX files
    if process_xlsx:
        xlsx_count = process_xlsx_files(input_dir, csv_output_dir, json_output_dir, TARGET_XLSX_FILES)
        total_processed += xlsx_count
    
    # Summary
    print("\n" + "="*60)
    print("Extraction Complete!")
    print("="*60)
    print(f"Total files processed: {total_processed}")
    print(f"Output saved to: {output_base_dir}")
    print("="*60)
    
    if total_processed == 0:
        print("\nWARNING: No files were processed. Check that target files exist in input directory.")
        sys.exit(1)


if __name__ == "__main__":
    main()
